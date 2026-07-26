# Excel Data Extractor.

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("personal_info.xlsx")
sheet = workbook.active

# Store data.
people = []

# Skip the header row.
for row in range(2, sheet.max_row + 1):
    person = {
        "Name": sheet.cell(row=row, column=1).value,
        "Age": sheet.cell(row=row, column=2).value,
        "City": sheet.cell(row=row, column=3).value
    }

    people.append(person)

# Display extracted data.
print("People Information:\n")

for person in people:
    print(f"Name: {person['Name']}")
    print(f"Age: {person['Age']}")
    print(f"City: {person['City']}")
    print("-" * 25)
    