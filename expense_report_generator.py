# Expense Report Generator.

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("expenses.xlsx")
sheet = workbook.active

# Store category totals.
expenses = {}

# Read expense records.
for row in range(2, sheet.max_row + 1):
    category = sheet.cell(row=row, column=1).value
    amount  = sheet.cell(row=row, column=2).value

    if category in expenses:
        expenses[category] += amount
    else:
        expenses[category] = amount

# Create summary sheet.
summary_sheet = workbook.create_sheet(title="Summary")

summary_sheet["A1"] = "Category"
summary_sheet["B1"] = "Total"

row = 2

for category, total in expenses.items():
    summary_sheet.cell(row=row, column=1).value = category
    summary_sheet.cell(row=row, column=2).value = total
    row += 1

# Save workbook.
workbook.save("expenses.xlsx")

# Display results.
print("Expense Summary")
print("-" * 25)

for category, total in expenses.items():
    print(f"{category}: {total}")

print("\nSummary sheet created successfully!")