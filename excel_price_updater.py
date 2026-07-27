# Excel Price Updater.

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("products.xlsx")
sheet = workbook.active

# Updated prices.
new_prices = {
    "Laptop": 850,
    "Mouse": 25,
    "Keyboard": 45,
    "Monitor": 170
}

# Update prices.
for row in range(2, sheet.max_row + 1):
    product = sheet.cell(row=row, column=1).value

    if product in new_prices:
        sheet.cell(row=row, column=2).value = new_prices[product]


# Save as a new workbook.
workbook.save("updated_products.xlsx")

print("Prices updated successfully!")
print("New file created: updated_products.xlsx")