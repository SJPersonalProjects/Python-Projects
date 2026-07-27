# Excel Inventory Tracker

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("inventory.xlsx")
sheet = workbook.active

# Store inventory.
inventory = {}

# Read spreadsheet.
for row in range(2, sheet.max_row + 1):
    product = sheet.cell(row=row, column=1).value
    quantity = sheet.cell(row=row, column=2).value

    inventory[product] = quantity

# Find low-stock items.
low_stock = {}

for product, quantity in inventory.items():
    if quantity < 5:
        low_stock[product] = quantity

# Create report.
with open("low_stock_report.txt", "w") as report:
    report.write("LOW STOCK REPORT\n")
    report.write("=" * 25 + "\n\n")

    for product, quantity in low_stock.items():
        report.write(f"{product}: {quantity} left\n")


# Display inventory.
print("Inventory")
print("-" * 20)

for product, quantity in inventory.items():
    print(f"{product}: {quantity}")

print("\nLow-stock report saved as 'low_stock_report.txt'.")