# Automated Business Spreadsheet Manager.

import openpyxl
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

# Files to combine
files = ["store_a.xlsx", "store_b.xlsx"]

# Create report workbook.
report_book = openpyxl.Workbook()
report_sheet = report_book.active
report_sheet.title = "Inventory Report"

# Title
report_sheet["A1"] = "Business Inventory Report"
report_sheet["A1"].font = Font(size=16, bold=True)

# Headers.
headers = ["Product", "Category", "Quantity", "Price", "Inventory Value"]

for column, header in enumerate(headers, start=1):
    cell = report_sheet.cell(row=3, column=column)
    cell.value = header
    cell.font = Font(bold=True)

current_row = 4

# Category totals.
category_totals = {}

# Read every workbook.
for file in files:

    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):

        product = sheet.cell(row=row, column=1).value
        category = sheet.cell(row=row, column=2).value
        quantity = sheet.cell(row=row, column=3).value
        price = sheet.cell(row=row, column=4).value

        inventory_value = quantity * price

        # Copy into report.
        report_sheet.cell(row=current_row, column=1).value = product
        report_sheet.cell(row=current_row, column=2).value = category
        report_sheet.cell(row=current_row, column=3).value = quantity
        report_sheet.cell(row=current_row, column=4).value = price
        report_sheet.cell(row=current_row, column=5).value = inventory_value

        # Update summary totals.
        if category in category_totals:
            category_totals[category] += inventory_value
        else:
            category_totals[category] = inventory_value
        
        current_row += 1

# Formatting.
for column in ['A', 'B', 'C', 'D', 'E']:
    report_sheet.column_dimensions[column].width = 18

report_sheet.freeze_panes = "A4"

# Summary sheet.
summary_sheet = report_book.create_sheet(title="Summary")

summary_sheet["A1"] = "Category"
summary_sheet["B1"] = "Total Inventory Value"

summary_sheet["A1"].font =  Font(bold=True)
summary_sheet["B1"].font = Font(bold=True)

summary_row = 2

for category, total in category_totals.items():

    summary_sheet.cell(row=summary_row, column=1).value = category
    summary_sheet.cell(row=summary_row, column=2).value = total

    summary_row += 1

# Create chart.
data = Reference(
    summary_sheet,
    min_col=2,
    min_row=1,
    max_row=summary_sheet.max_row
)

categories = Reference(
    summary_sheet,
    min_col=1,
    min_row=2,
    max_row=summary_sheet.max_row
)

chart = BarChart()
chart.title = "Inventory Value by Category"
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

summary_sheet.add_chart(chart, "D2")


# Save report.
report_book.save("business_report.xlsx")

print("Business report generated successfully!")
print("Output file: business_report.xlsx")