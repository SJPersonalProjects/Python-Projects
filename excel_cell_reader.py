# Excel Cell Reader.

import openpyxl

# Load the Excel Workbook.
workbook = openpyxl.load_workbook("sample.xlsx")

# Selected the active sheet.
sheet = workbook.active

# Read the specific cells.
print("A1:", sheet["A1"].value)
print("B2:", sheet["B2"].value)
print("C3:", sheet["C3"].value)

# Another way to access cells.
print("\nUsing row and column numbers.")
print("A1:", sheet.cell(row=1, column=1).value)
print("B2:", sheet.cell(row=2, column=2).value)
print("C3:", sheet.cell(row=3, column=3).value)