# Spreadsheet data cleaner.

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("employees.xlsx")
sheet = workbook.active

# Scan every data row.
for row in range(2, sheet.max_row + 1):

    # Check Name.
    if sheet.cell(row=row, column=1).value is None:
        sheet.cell(row=row, column=1).value = "Missing"

    # Check Age.
    age = sheet.cell(row=row, column=2).value

    if age is None:
        sheet.cell(row=row, column=2).value = "Missing"
    elif age < 0:
        sheet.cell(row=row, column=2).value = 0

    # Check Department.
    if sheet.cell(row=row, column=3).value is None:
        sheet.cell(row=row, column=3).value = "Missing"

# Save changes.
workbook.save("employees.xlsx")

print("Spreadsheet cleaned successfully!")