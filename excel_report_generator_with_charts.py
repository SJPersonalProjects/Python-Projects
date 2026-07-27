# Excel Report Generator with Charts.

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font

# Load workbook.
workbook = openpyxl.load_workbook("sales.xlsx")
sheet = workbook.active

# Create report sheet.
report_sheet = workbook.create_sheet(title="Sales Report")

# Title.
report_sheet["A1"] = "Monthly Sales Report"
report_sheet["A1"].font = Font(size=16, bold=True)

# Headers.
report_sheet["A3"] = "Month"
report_sheet["B3"] = "Sales"

report_sheet["A3"].font = Font(bold=True)
report_sheet["B3"].font = Font(bold=True)

# Copy data from the original sheet
report_row = 4

for row in range(2, sheet.max_row + 1):
    month = sheet.cell(row=row, column=1).value
    sales = sheet.cell(row=row, column=2).value

    report_sheet.cell(row=report_row, column=1).value = month
    report_sheet.cell(row=report_row, column=2).value = sales

    report_row += 1

# References for charts.
data = Reference(report_sheet,
                 min_col=2,
                 min_row=3,
                 max_row=report_sheet.max_row)

categories = Reference(report_sheet,
                       min_col=1,
                       min_row=4,
                       max_row=report_sheet.max_row)


# Bar Chart.
bar_chart = BarChart()
bar_chart.title = "Monthly Sales"
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)

# Line Chart.
line_chart = LineChart()
line_chart.title = "Sales Trend"
line_chart.add_data(data, titles_from_data=True)
line_chart.set_categories(categories)

# Pie Chart
pie_chart = PieChart()
pie_chart.title = "Sales Distribution"
pie_chart.add_data(data, titles_from_data=True)
pie_chart.set_categories(categories)

# Add Charts.
report_sheet.add_chart(bar_chart, "D2")
report_sheet.add_chart(line_chart, "D18")
report_sheet.add_chart(pie_chart, "D34")

# Save workbook.
workbook.save("sales_report.xlsx")

print("Sales report created successfully!")
print("Saved as 'sales_report.xlsx")