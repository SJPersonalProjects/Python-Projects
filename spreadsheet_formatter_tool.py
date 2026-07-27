# Spreadsheet formatter tool.

import openpyxl
from openpyxl.styles import Font

# Load workbook.
workbook = openpyxl.load_workbook("students_marks.xlsx")
sheet = workbook.active

# Insert title row.
sheet.insert_rows(1)

# Marge title cells.
sheet.merge_cells("A1:C1")
sheet["A1"] = "Student Marks Report"

# Format title.
sheet["A1"].font = Font(name="Calibri", size=16, bold=True)

# Make header row bold.
for column in range(1, sheet.max_column + 1):
    sheet.cell(row=2, column=column).font = Font(bold=True)

# Format data.
for row in range(3, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        sheet.cell(row=row, column=column).font = Font(name="calibri", size=11)


# Adjust column widths
sheet.column_dimensions["A"].width = 20
sheet.column_dimensions["B"].width = 12
sheet.column_dimensions["C"].width = 12

# Freeze title and header rows.
sheet.freeze_panes = "A3"

# Save workbook.
workbook.save("formatted_students_marks.xlsx")

print("Spreadsheet formatted successfully!")
print("Saved as 'formatted_student_marks.xlsx")