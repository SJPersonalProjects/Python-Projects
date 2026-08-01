# Unpaid Invoices Reminder.

import openpyxl
import smtplib

# Load Excel file.
workbook = openpyxl.load_workbook("invoices.xlsx")
sheet = workbook.active

# Email details.
sender_email = input("Your email: ")
password = input("Your app password: ")

# Connect to SMTP
smtp = smtplib.SMTP("smtp.gmail.com", 587)

smtp.ehlo()
smtp.starttls()
smtp.ehlo()

smtp.login(
    sender_email,
    password
)

# Read customers.
for row in range(2, sheet.max_row + 1):

    name = sheet.cell(row, 1).value
    email = sheet.cell(row, 2).value
    amount = sheet.cell(row, 3).value
    status = sheet.cell(row, 4).value

    # Check unpaid invoices.
    if status.lower() == "unpaid":

        subject = "Payment Reminder"
        message = f"""Subject: {subject}

Dear {name},

This is a friendly reminder that your invoices
of ${amount} is still unpaid.

Please complete your payment at your earliest convenience.

Thank you.
"""
        
    smtp.sendmail(
        sender_email,
        email,
        message
    )

    print(f"Reminder sent to {name}")

smtp.quit()

print("All reminders completed.")