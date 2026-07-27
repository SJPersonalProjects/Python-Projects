# Spreadsheet Satistics Calculator

import openpyxl

# Load workbook.
workbook = openpyxl.load_workbook("scores.xlsx")
sheet = workbook.active

# Read numbers from column A and skip header.
numbers = []

for row in range(2, sheet.max_row + 1):
    value = sheet.cell(row=row, column=1).value
    numbers.append(value)


# Calculate statistics
total = sum(numbers)
average = total / len(numbers)
maximum = max(numbers)
minimum = min(numbers)

# Create a new sheet.
stats_sheet = workbook.create_sheet(title="Statistics")

# Write results.
stats_sheet["A1"] = "Statistics"
stats_sheet["B1"] = "Value"

stats_sheet["A2"] = "Total"
stats_sheet["B2"] = total

stats_sheet["A3"] = "Average"
stats_sheet["B3"] = average

stats_sheet["A4"] = "Maximum"
stats_sheet["B4"] = maximum

stats_sheet["A5"] = "Minimum"
stats_sheet["B5"] = minimum

# Save the workbook.
workbook.save("scores.xlsx")

# Display results.
print("Statistics")
print("-" * 20)
print("Total    :", total)
print("Average  :", average)
print("Maximum  :", maximum)
print("Minimum  :", minimum)
print("\nResults saved in the 'Statistics' sheet.")